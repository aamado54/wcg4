#!/usr/bin/env python3
"""wcup2 — consolidación de estados financieros WCG (evolución de wcup1).

Pasos:
  1) Renombrar archivos legacy en el directorio fuente al canónico
     WCF|WCL|WCI|WCS - ER|BG - yyMM.xlsx
  2) Consolidar contra la estructura (chart of accounts) y emitir Excel.

Config (wconfig12.txt, 3 líneas):
  1. archivo de estructura (.xlsx)
  2. ruta del directorio de entrada (wcsource)
  3. archivo de salida (.xlsx)

Uso:
  python wcup2.py
  python wcup2.py --skip-rename
  python wcup2.py --rename-only
"""

from __future__ import annotations

import argparse
import glob
import os
import re
import sqlite3
import sys
from pathlib import Path

import openpyxl

CANON_RE = re.compile(r"^(WCF|WCL|WCI|WCS)-(ER|BG)-(\d{4})\.xlsx$", re.I)
LEGACY_RE = re.compile(
    r"^(?:\d+\.\s*)?(WC(?:F|L|I|S)?)\s*-\s*"
    r"(Estado de resultados|Balance general)_"
    r"(enero|febrero|marzo|abril|mayo|junio|julio|agosto|"
    r"septiembre|setiembre|octubre|noviembre|diciembre)\s+(\d{4})\.xlsx$",
    re.I,
)
MONTHS = {
    "enero": "01",
    "febrero": "02",
    "marzo": "03",
    "abril": "04",
    "mayo": "05",
    "junio": "06",
    "julio": "07",
    "agosto": "08",
    "septiembre": "09",
    "setiembre": "09",
    "octubre": "10",
    "noviembre": "11",
    "diciembre": "12",
}
VALID_ENTS = {"WCF", "WCL", "WCI", "WCS"}


def _norm_header(v) -> str:
    if v is None:
        return ""
    s = str(v).strip().upper()
    s = s.replace("Á", "A").replace("É", "E").replace("Í", "I").replace("Ó", "O").replace("Ú", "U")
    s = re.sub(r"\s+", " ", s)
    return s


def detect_columns(headers: list) -> tuple[int, int] | None:
    """Detect (acode_col, aname_col) from row-1 headers. Value = last numeric cell."""
    h = [_norm_header(x) for x in headers]
    # Prefer explicit account-number headers
    for code_name in ("NUMERO CUENTA", "NRO CUENTA", "NO. CUENTA", "NO CUENTA"):
        if code_name in h:
            ci = h.index(code_name)
            # Name often labeled CUENTA or NOMBRE
            for name_label in ("NOMBRE", "CUENTA", "DESCRIPCION"):
                if name_label in h and h.index(name_label) != ci:
                    return ci, h.index(name_label)
            # fallback: next text column
            return ci, ci + 1 if ci + 1 < len(h) else ci
    if "CUENTA" in h and "NOMBRE" in h:
        return h.index("CUENTA"), h.index("NOMBRE")
    if "CUENTA" in h:
        ci = h.index("CUENTA")
        # Type1 legacy: CUENTA | NOMBRE | SALDOFIN
        if ci + 1 < len(h):
            return ci, ci + 1
    return None


def last_numeric(row) -> float:
    numval = None
    for i in range(len(row) - 1, -1, -1):
        cell = row[i]
        if cell is None or cell == "":
            continue
        try:
            numval = float(cell)
            break
        except (TypeError, ValueError):
            continue
    return 0.0 if numval is None else numval


def rename_source_files(input_dir: str) -> dict:
    """Rename legacy Spanish filenames to WCx-ER/BG-yyMM.xlsx. Returns summary."""
    src = Path(os.path.expanduser(input_dir))
    summary = {"renamed": [], "skipped": [], "errors": []}
    if not src.is_dir():
        summary["errors"].append(f"input_dir missing: {src}")
        return summary

    for p in sorted(src.iterdir()):
        if not p.is_file() or p.suffix.lower() != ".xlsx":
            continue
        name = p.name
        if CANON_RE.match(name):
            summary["skipped"].append((name, "already_canonical"))
            continue
        m = LEGACY_RE.match(name)
        if not m:
            summary["errors"].append((name, "unmatched_pattern"))
            continue
        raw_ent, stmt, month_es, year = m.group(1), m.group(2), m.group(3), m.group(4)
        ent = raw_ent.upper()
        if ent == "WC":
            ent = "WCF"
        if ent not in VALID_ENTS:
            summary["errors"].append((name, f"bad_entity:{ent}"))
            continue
        kind = "ER" if stmt.lower() == "estado de resultados" else "BG"
        new_name = f"{ent}-{kind}-{year[-2:]}{MONTHS[month_es.lower()]}.xlsx"
        dest = src / new_name
        if dest.exists() and dest.resolve() != p.resolve():
            summary["errors"].append((name, f"collision:{new_name}"))
            continue
        p.rename(dest)
        summary["renamed"].append((name, new_name))
    return summary


def parse_filename(file_name: str) -> tuple[str, int] | None:
    """Extract (businessid single char, perid YYMM) from canonical name."""
    m = CANON_RE.match(file_name)
    if m:
        return m.group(1)[2].upper(), int(m.group(3))
    # Fallback legacy: 3rd char + last 4 digits
    try:
        businessid = file_name[2].upper().strip()
        period_matches = re.findall(r"\d{4}", file_name)
        if not period_matches:
            return None
        return businessid, int(period_matches[-1])
    except (IndexError, ValueError):
        return None


def load_config(path: str = "wconfig12.txt") -> tuple[str, str, str]:
    with open(path, "r", encoding="utf-8") as f:
        lines = [line.strip() for line in f.readlines() if line.strip()]
    if len(lines) < 3:
        raise SystemExit(f"Config {path} needs 3 lines: structure, input_dir, output")
    structure = lines[0] if lines[0].lower().endswith(".xlsx") else lines[0] + ".xlsx"
    input_dir = os.path.expanduser(lines[1])
    output = lines[2] if lines[2].lower().endswith(".xlsx") else lines[2] + ".xlsx"
    return structure, input_dir, output


def consolidate(structurefile: str, input_dir: str, output_file: str) -> dict:
    conn = sqlite3.connect(":memory:")
    cursor = conn.cursor()
    cursor.execute(
        "CREATE TABLE blanks (id INTEGER PRIMARY KEY AUTOINCREMENT, ln INTEGER)"
    )
    cursor.execute(
        """CREATE TABLE alines (
        aid INTEGER PRIMARY KEY AUTOINCREMENT,
        businessid TEXT, acode TEXT, aname TEXT, newcode INTEGER)"""
    )
    cursor.execute(
        """CREATE TABLE numvals (
        nid INTEGER PRIMARY KEY AUTOINCREMENT,
        aid INTEGER, perid INTEGER, numval REAL)"""
    )
    cursor.execute("CREATE TABLE pers (persid INTEGER)")
    conn.commit()

    wb = openpyxl.load_workbook(structurefile)
    if len(wb.sheetnames) != 1:
        print(f"Error: Structure file contains {len(wb.sheetnames)} sheets")
        return {"ok": False, "error": "structure_sheets"}
    sheet = wb.active
    line_counter = 2
    for row in sheet.iter_rows(min_row=2, values_only=True):
        businessid = str(row[1]).strip() if row[1] is not None else ""
        acode = str(row[2]).strip() if row[2] is not None else ""
        acode = re.sub(r"\.0$", "", acode)
        if not acode:
            cursor.execute("INSERT INTO blanks (ln) VALUES (?)", (line_counter,))
            line_counter += 1
            continue
        aname = str(row[3]).strip() if row[3] is not None else ""
        cursor.execute(
            "INSERT INTO alines (businessid, acode, aname, newcode) VALUES (?, ?, ?, 0)",
            (businessid, acode, aname),
        )
        line_counter += 1
    conn.commit()

    files = glob.glob(os.path.join(input_dir, "*.xlsx"))
    total_files = len(files)
    processed = 0
    skipped = []
    format_hits: dict[str, int] = {}

    def sort_key(file_path: str) -> int:
        parsed = parse_filename(os.path.basename(file_path))
        return parsed[1] if parsed else 0

    print(f"\nFiles processed: 0/{total_files}", end="", flush=True)
    for file_path in sorted(files, key=sort_key):
        file_name = os.path.basename(file_path)
        print(f"\rFiles processed: {processed + 1}/{total_files}", end="", flush=True)
        parsed = parse_filename(file_name)
        if not parsed:
            skipped.append((file_name, "bad_name"))
            processed += 1
            continue
        businessid, perid = parsed
        cursor.execute("INSERT INTO pers (persid) VALUES (?)", (perid,))
        try:
            wb_in = openpyxl.load_workbook(file_path, data_only=True)
            sheet_in = wb_in["Datos"]
        except Exception as exc:  # noqa: BLE001
            skipped.append((file_name, f"open:{exc}"))
            processed += 1
            continue

        header_row = next(sheet_in.iter_rows(min_row=1, max_row=1, values_only=True))
        cols = detect_columns(list(header_row))
        if cols is None:
            skipped.append((file_name, f"headers:{header_row[:6]}"))
            processed += 1
            continue
        acode_col, aname_col = cols
        fmt_key = f"acode={acode_col},aname={aname_col}"
        format_hits[fmt_key] = format_hits.get(fmt_key, 0) + 1

        for row in sheet_in.iter_rows(min_row=2, values_only=True):
            if acode_col >= len(row):
                continue
            acode = str(row[acode_col]).strip() if row[acode_col] is not None else ""
            acode = re.sub(r"\.0$", "", acode)
            if not acode or acode.upper() in ("NONE", "NULL"):
                continue
            aname = ""
            if aname_col < len(row) and row[aname_col] is not None:
                aname = str(row[aname_col]).strip()
            numval = last_numeric(row)

            cursor.execute(
                """SELECT aid FROM alines
                   WHERE businessid = ? AND TRIM(LOWER(acode)) = LOWER(?)""",
                (businessid, acode),
            )
            result = cursor.fetchone()
            if result:
                aid = result[0]
            else:
                cursor.execute(
                    """INSERT INTO alines (businessid, acode, aname, newcode)
                       VALUES (?, ?, ?, 1)""",
                    (businessid, acode, aname),
                )
                aid = cursor.lastrowid
            # Upsert-like: replace prior value for same aid+period (BG/ER collision rare)
            cursor.execute(
                "DELETE FROM numvals WHERE aid = ? AND perid = ?", (aid, perid)
            )
            cursor.execute(
                "INSERT INTO numvals (aid, perid, numval) VALUES (?, ?, ?)",
                (aid, perid, numval),
            )
        conn.commit()
        processed += 1

    cursor.execute("SELECT DISTINCT persid FROM pers ORDER BY persid")
    persids = [row[0] for row in cursor.fetchall()]
    print(f"\nDetected periods: {persids}")
    print(f"Format hits: {format_hits}")
    if skipped:
        print(f"Skipped ({len(skipped)}):")
        for item in skipped[:20]:
            print(" ", item)

    cursor.execute("DROP TABLE IF EXISTS outputt")
    cursor.execute("CREATE TABLE outputt AS SELECT * FROM alines")
    for pid in persids:
        try:
            cursor.execute(f"ALTER TABLE outputt ADD COLUMN n{pid} REAL")
        except sqlite3.OperationalError:
            pass

    cursor.execute("SELECT aid FROM outputt")
    aids = [row[0] for row in cursor.fetchall()]
    for aid in aids:
        for pid in persids:
            cursor.execute(
                "SELECT numval FROM numvals WHERE aid = ? AND perid = ?", (aid, pid)
            )
            result = cursor.fetchone()
            if result:
                cursor.execute(
                    f"UPDATE outputt SET n{pid} = ? WHERE aid = ?",
                    (result[0], aid),
                )
    conn.commit()

    cursor.execute("SELECT ln FROM blanks")
    blank_lines = {row[0] for row in cursor.fetchall()}
    cursor.execute("SELECT * FROM outputt")
    output_rows = cursor.fetchall()
    cursor.execute("PRAGMA table_info(outputt)")
    columns = [col[1] for col in cursor.fetchall()]
    header = columns.copy()
    header[0] = ""

    rows_0 = [row for row in output_rows if row[4] == 0]
    rows_1 = [row for row in output_rows if row[4] == 1]
    business_order = {"F": 0, "L": 1, "I": 2, "S": 3}
    rows_1 = sorted(rows_1, key=lambda x: (business_order.get(x[1], 99), str(x[2])))

    wb_out = openpyxl.Workbook()
    ws = wb_out.active
    ws.title = "consolidado"
    ws.append(header)

    current_input_line = 2
    for row in rows_0:
        while current_input_line in blank_lines:
            ws.append([""] * len(header))
            current_input_line += 1
        row_list = list(row)
        row_list[0] = None
        ws.append(row_list)
        current_input_line += 1

    ws.append([""] * len(header))
    ws.append([""] * len(header))

    prev_business = None
    for row in rows_1:
        if prev_business is not None and row[1] != prev_business:
            ws.append([""] * len(header))
        row_list = list(row)
        row_list[0] = None
        ws.append(row_list)
        prev_business = row[1]

    for col_idx in range(5, len(columns)):
        col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
        for cell in ws[col_letter]:
            if cell.row >= 2 and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"

    # Meta sheet for review
    meta = wb_out.create_sheet("meta", 0)
    meta.append(["campo", "valor"])
    meta.append(["structure", structurefile])
    meta.append(["input_dir", input_dir])
    meta.append(["output", output_file])
    meta.append(["files_total", total_files])
    meta.append(["files_processed", processed - len(skipped)])
    meta.append(["files_skipped", len(skipped)])
    meta.append(["periods", ", ".join(str(p) for p in persids)])
    meta.append(["format_hits", str(format_hits)])
    for name, reason in skipped:
        meta.append(["skip", f"{name} :: {reason}"])

    wb_out.save(output_file)
    conn.close()
    print(f"\nProcess completed. Output saved to: {output_file}")
    return {
        "ok": True,
        "periods": persids,
        "files_total": total_files,
        "skipped": skipped,
        "format_hits": format_hits,
        "output": output_file,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="wcup2 WCG financial consolidator")
    parser.add_argument("--config", default="wconfig12.txt")
    parser.add_argument("--skip-rename", action="store_true")
    parser.add_argument("--rename-only", action="store_true")
    args = parser.parse_args(argv)

    structure, input_dir, output = load_config(args.config)

    if not args.skip_rename:
        print(f"Renaming files in {input_dir} …")
        summary = rename_source_files(input_dir)
        print(f"  renamed={len(summary['renamed'])} skipped={len(summary['skipped'])} errors={len(summary['errors'])}")
        for a, b in summary["renamed"][:10]:
            print(f"    {a} -> {b}")
        if len(summary["renamed"]) > 10:
            print(f"    … +{len(summary['renamed']) - 10} more")
        for err in summary["errors"]:
            print(f"  ERROR {err}")
        if args.rename_only:
            return 0 if not summary["errors"] else 1

    consolidate(structure, input_dir, output)
    return 0


if __name__ == "__main__":
    sys.exit(main())
