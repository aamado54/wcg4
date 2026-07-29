#!/usr/bin/env python3
"""Regenera combined_series.json a partir de wcout2d + wc-mod5c.

Uso (desde raíz del repo dashboard):
  .venv/bin/python tools/financiero/rebuild_combined.py
"""

from __future__ import annotations

import json
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
BASE = ROOT / "data" / "wcg" / "financiero"
BU_LABEL = {
    "F": "Factoraje",
    "L": "Leasing",
    "I": "Insurance",
    "S": "Services",
    "T": "Total",
}
SUMMARY_CODES = {
    "activo": ["1"],
    "activo_corriente": ["101"],
    "activo_no_corriente": ["102"],
    "pasivo_corriente": ["201"],
    "pasivo_no_corriente": ["202"],
    "patrimonio": ["301"],
    "utilidades": ["302"],
    "cartera": ["1010301", "10103"],
}


def main() -> int:
    hist_path = BASE / "wc-mod5c.xlsx"
    recent_path = BASE / "wcout2d.xlsx"
    if not hist_path.exists() or not recent_path.exists():
        print("Missing inputs under", BASE, file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(hist_path, data_only=True, read_only=True)
    ws = wb["Datos"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header = rows[0]
    col_to_period = {}
    for ci, v in enumerate(header):
        if ci < 12 or v is None:
            continue
        if isinstance(v, datetime):
            p = f"{v.year}-{v.month:02d}"
        else:
            m = re.match(
                r"(Ene|Feb|Mar|Abr|May|Jun|Jul|Ago|Sep|Oct|Nov|Dic)\s*(\d{4})",
                str(v).strip(),
                re.I,
            )
            if not m:
                continue
            mm = {
                "ene": "01",
                "feb": "02",
                "mar": "03",
                "abr": "04",
                "may": "05",
                "jun": "06",
                "jul": "07",
                "ago": "08",
                "sep": "09",
                "oct": "10",
                "nov": "11",
                "dic": "12",
            }[m.group(1).lower()[:3]]
            p = f"{m.group(2)}-{mm}"
        col_to_period[ci] = p

    accounts = []
    for row in rows[12:]:
        if not row or row[1] is None:
            continue
        code = re.sub(r"\.0$", "", str(row[1]).strip())
        if not code or not re.match(r"^\d", code):
            continue
        bu = str(row[2]).strip().upper() if row[2] else ""
        if bu not in ("F", "L", "I", "S"):
            continue
        label = str(row[3]).strip() if row[3] else ""
        vals = {
            p: float(row[ci])
            for ci, p in col_to_period.items()
            if ci < len(row) and isinstance(row[ci], (int, float))
        }
        if vals:
            accounts.append(
                {
                    "bu": bu,
                    "code": code,
                    "label": label,
                    "values": vals,
                    "source": "historico",
                }
            )

    wb = openpyxl.load_workbook(recent_path, data_only=True)
    ws = wb["consolidado"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    per_cols = {}
    for i, h in enumerate(headers):
        if h and str(h).startswith("n") and str(h)[1:].isdigit():
            yymm = str(h)[1:]
            per_cols[i] = f"{2000 + int(yymm[:2])}-{int(yymm[2:]):02d}"

    recent_map = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1] or not row[2]:
            continue
        bu = str(row[1]).strip().upper()
        code = re.sub(r"\.0$", "", str(row[2]).strip())
        if bu not in ("F", "L", "I", "S"):
            continue
        key = (bu, code)
        if key not in recent_map:
            recent_map[key] = {
                "bu": bu,
                "code": code,
                "label": str(row[3]).strip() if row[3] else "",
                "values": {},
                "source": "import_reciente",
            }
        for ci, p in per_cols.items():
            if ci < len(row) and isinstance(row[ci], (int, float)):
                recent_map[key]["values"][p] = float(row[ci]) / 1000.0
    wb.close()

    by_key = {(a["bu"], a["code"]): dict(a, values=dict(a["values"])) for a in accounts}
    for key, rec in recent_map.items():
        if key in by_key:
            by_key[key]["values"].update(rec["values"])
            by_key[key]["source"] = "combinado"
        else:
            by_key[key] = rec

    merged = list(by_key.values())
    all_periods = sorted({p for a in merged for p in a["values"]})

    def get_val(bu, codes, period):
        for c in codes:
            a = by_key.get((bu, c))
            if a and period in a["values"]:
                return a["values"][period]
        return None

    kpis = {bu: {} for bu in "FLIS"}
    for bu in "FLIS":
        for p in all_periods:
            row = {
                name: get_val(bu, codes, p)
                for name, codes in SUMMARY_CODES.items()
                if get_val(bu, codes, p) is not None
            }
            # rebuild properly
            row = {}
            for name, codes in SUMMARY_CODES.items():
                v = get_val(bu, codes, p)
                if v is not None:
                    row[name] = v
            ac, pc = row.get("activo_corriente"), row.get("pasivo_corriente")
            if ac is not None and pc:
                row["liquidez"] = ac / pc
            act, util = row.get("activo"), row.get("utilidades")
            if act and util is not None:
                row["roa"] = util / act
            pat = row.get("patrimonio")
            if pat and util is not None:
                row["roe"] = util / pat
            pas = (row.get("pasivo_corriente") or 0) + (row.get("pasivo_no_corriente") or 0)
            if pat:
                row["apalancamiento"] = pas / pat
            if row:
                kpis[bu][p] = row

    kpis["T"] = {}
    for p in all_periods:
        agg = {}
        for bu in "FLIS":
            for k, v in (kpis[bu].get(p) or {}).items():
                if k in ("liquidez", "roa", "roe", "apalancamiento"):
                    continue
                if isinstance(v, (int, float)):
                    agg[k] = agg.get(k, 0.0) + v
        if agg.get("activo_corriente") and agg.get("pasivo_corriente"):
            agg["liquidez"] = agg["activo_corriente"] / agg["pasivo_corriente"]
        if agg.get("activo") and agg.get("utilidades") is not None:
            agg["roa"] = agg["utilidades"] / agg["activo"]
        pat = agg.get("patrimonio")
        if pat and agg.get("utilidades") is not None:
            agg["roe"] = agg["utilidades"] / pat
        pas = (agg.get("pasivo_corriente") or 0) + (agg.get("pasivo_no_corriente") or 0)
        if pat:
            agg["apalancamiento"] = pas / pat
        if agg:
            kpis["T"][p] = agg

    qf_path = BASE / "wc_mod5c_extract.json"
    qf = json.loads(qf_path.read_text()) if qf_path.exists() else {}

    payload = {
        "unit": "000 quetzales",
        "generated": datetime.now().isoformat(timespec="seconds"),
        "history_source": "wc-mod5c.xlsx#Datos",
        "recent_source": "wcout2d.xlsx (estados financieros /1000)",
        "periods": all_periods,
        "recent_periods": sorted(set(per_cols.values())),
        "business_units": BU_LABEL,
        "kpis": kpis,
        "accounts_count": len(merged),
        "qf_extract_meta": {
            "history_periods": qf.get("history_periods"),
            "forecast_periods": qf.get("forecast_periods"),
            "accounts": len(qf.get("accounts") or []),
            "control_defaults": qf.get("control_defaults"),
        },
        "accounts_sample": [
            a
            for a in merged
            if a["code"]
            in {"1", "101", "102", "2", "201", "202", "301", "302", "4", "5", "6", "1010301"}
            or (a["code"].startswith(("4", "5", "6", "7", "8")) and len(a["code"]) <= 3)
        ],
    }
    (BASE / "combined_series.json").write_text(
        json.dumps(payload, ensure_ascii=False), encoding="utf-8"
    )
    full = {**payload, "accounts": merged}
    (BASE / "combined_full.json").write_text(
        json.dumps(full, ensure_ascii=False), encoding="utf-8"
    )
    print(
        f"OK periods={len(all_periods)} accounts={len(merged)} → {BASE/'combined_series.json'}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
