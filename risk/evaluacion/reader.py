"""Lectura segura de la plantilla de evaluación de riesgo de clientes.

Interfaz estable para importación formal (no cambiar firma sin migrar callers):

    load_evaluacion()                         # settings / docs default
    load_evaluacion(path="/ruta/archivo.xlsx")
    load_evaluacion(uploaded_file=file_obj)   # Django UploadedFile / BinaryIO

Siempre retorna `EvaluacionDataset` (nunca lanza hacia la vista). Errores van
en `errors`; si no hay empresas, `companies == []` y la UI muestra vacío limpio.
"""

from __future__ import annotations

import logging
import re
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, BinaryIO

from django.conf import settings

logger = logging.getLogger(__name__)

SHEETS_REQUIRED = ("Caratula", "Altman")
SHEETS_OPTIONAL = ("Balance", "Resultados")

# Modelo Z preferido en Carátula (fila de calificación Bien/Soso/Mal).
# Clave interna: z_emergentes (modelo 3).

CARATULA_METRIC_ALIASES: dict[str, str] = {
    "ventas": "ventas",
    "utilidad neta": "utilidad_neta",
    "activo corriente": "activo_corriente",
    "activo no corriente": "activo_no_corriente",
    "total de activo": "total_activo",
    "pasivo corriente": "pasivo_corriente",
    "pasivo no corriente": "pasivo_no_corriente",
    "total pasivo": "total_pasivo",
    "total patrimonio": "total_patrimonio",
    "liquidez": "liquidez",
    "prueba acida": "prueba_acida",
    "prueba ácida": "prueba_acida",
    "capital de trabajo": "capital_trabajo",
    "ebitda": "ebitda",
    "apalancamiento": "apalancamiento",
    "crecimiento anual en ventas": "crecimiento_ventas",
    "roe": "roe",
    "roa": "roa",
    "ciclo de caja en días": "ciclo_caja_dias",
    "período promedio de cobro (días)": "dias_cobro",
    "periodo promedio de cobro (días)": "dias_cobro",
}

Z_MODEL_ALIASES: dict[str, str] = {
    "1. altman original": "z_altman_1968",
    "2. altman revisado": "z_altman_1983",
    "3. z'' para no manufactureras": "z_emergentes",
    "4. altman modificado": "z_modificado",
    "5. z’-score revisado": "z_pymes_ece",
    "5. z'-score revisado": "z_pymes_ece",
    "6. z''-score mercados emergentes": "z_emergentes_const",
    "7. pesos iguales": "z_pesos_iguales",
    "8. ajuste para empresas medianas": "z_latam",
}


@dataclass
class PeriodPoint:
    year: int
    status: str | None = None
    metrics: dict[str, float | None] = field(default_factory=dict)
    z_scores: dict[str, float | None] = field(default_factory=dict)
    z_label: str | None = None  # Bien / Soso / Mal


@dataclass
class CompanyEval:
    code: str
    name: str
    currency_unit: str | None = None
    giro: str | None = None
    auditor: str | None = None
    periods: list[PeriodPoint] = field(default_factory=list)

    def latest(self) -> PeriodPoint | None:
        if not self.periods:
            return None
        return max(self.periods, key=lambda p: p.year)


@dataclass
class AltmanModel:
    id: int
    name: str
    low_cut: float | None
    high_cut: float | None
    bajo_riesgo: str | None
    riesgo_moderado: str | None
    alto_riesgo: str | None


@dataclass
class EvaluacionDataset:
    source: str
    sheets_read: list[str]
    sheets_missing: list[str]
    companies: list[CompanyEval]
    altman_models: list[AltmanModel]
    primary_z_key: str = "z_emergentes"
    errors: list[str] = field(default_factory=list)

    @property
    def has_data(self) -> bool:
        return bool(self.companies)

    @property
    def status(self) -> str:
        """ok | empty | error — para UI sin inspeccionar listas."""
        if self.errors and not self.companies:
            return "error"
        if not self.companies:
            return "empty"
        if self.errors:
            return "partial"
        return "ok"

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def default_xlsx_path() -> Path:
    configured = getattr(settings, "WCG_EVALUACION_RIESGO_XLSX", None)
    if configured:
        return Path(configured)
    return Path(settings.BASE_DIR) / "docs" / "WCG-evaluacion-riesgo-clientes-2025.xlsx"


def _empty_dataset(source: str, errors: list[str]) -> EvaluacionDataset:
    return EvaluacionDataset(
        source=source or "(sin fuente)",
        sheets_read=[],
        sheets_missing=list(SHEETS_REQUIRED + SHEETS_OPTIONAL),
        companies=[],
        altman_models=[],
        errors=errors,
    )


def _resolve_sheet(wb, logical_name: str):
    """Match case-insensitive / espacios; tolera renombres leves."""
    wanted = logical_name.strip().lower()
    for name in wb.sheetnames:
        if name.strip().lower() == wanted:
            return name, wb[name]
    # Prefijo: "Caratula_V2" no sustituye Caratula, pero "Carátula " sí.
    for name in wb.sheetnames:
        if name.strip().lower().replace("á", "a") == wanted.replace("á", "a"):
            return name, wb[name]
    return None, None


def load_evaluacion(
    path: str | Path | None = None,
    *,
    uploaded_file: BinaryIO | None = None,
) -> EvaluacionDataset:
    """Carga la plantilla. Prioridad: uploaded_file > path > settings default.

    Nunca propaga excepciones: falla con EvaluacionDataset.status in {empty, error}.
    """
    try:
        import openpyxl
    except ImportError:  # pragma: no cover
        return _empty_dataset(
            "(sin openpyxl)",
            ["No se puede leer Excel: falta el paquete openpyxl."],
        )

    source_label = ""
    try:
        if uploaded_file is not None:
            uploaded_file.seek(0)
            wb = openpyxl.load_workbook(
                uploaded_file, data_only=True, read_only=True, keep_vba=False
            )
            source_label = getattr(uploaded_file, "name", "upload.xlsx")
        else:
            xlsx = Path(path) if path else default_xlsx_path()
            if not xlsx.is_file():
                return _empty_dataset(
                    str(xlsx),
                    [
                        "No hay plantilla de evaluación disponible. "
                        "Coloque el archivo en docs/ o configure WCG_EVALUACION_RIESGO_XLSX."
                    ],
                )
            wb = openpyxl.load_workbook(
                xlsx, data_only=True, read_only=True, keep_vba=False
            )
            source_label = str(xlsx)
    except Exception:
        logger.exception("No se pudo abrir plantilla de evaluación")
        return _empty_dataset(
            source_label or str(path or ""),
            ["No se pudo abrir el archivo Excel. Verifique que sea una plantilla .xlsx válida."],
        )

    errors: list[str] = []
    sheets_read: list[str] = []
    sheets_missing: list[str] = []
    sheet_map: dict[str, Any] = {}

    for name in SHEETS_REQUIRED + SHEETS_OPTIONAL:
        resolved, ws = _resolve_sheet(wb, name)
        if resolved is not None:
            sheets_read.append(name)
            sheet_map[name] = ws
        else:
            sheets_missing.append(name)
            if name in SHEETS_REQUIRED:
                errors.append(
                    f"Falta la hoja «{name}» (o fue renombrada). "
                    "Se necesita para el tablero gerencial."
                )

    companies: list[CompanyEval] = []
    altman_models: list[AltmanModel] = []

    try:
        if "Caratula" in sheet_map:
            try:
                companies = _parse_caratula(sheet_map["Caratula"], errors)
            except Exception:
                logger.exception("Parseo Caratula")
                errors.append("La hoja Carátula no se pudo interpretar por completo.")
        if "Altman" in sheet_map:
            try:
                altman_models = _parse_altman(sheet_map["Altman"], errors)
            except Exception:
                logger.exception("Parseo Altman")
                errors.append(
                    "La hoja Altman no se pudo interpretar; se usan umbrales por defecto."
                )
        _note_optional_sheets(sheet_map, sheets_read, errors)
    except Exception:
        logger.exception("Error parseando plantilla de evaluación")
        errors.append("Error inesperado al leer la plantilla.")
    finally:
        try:
            wb.close()
        except Exception:
            pass

    if not companies and not any("Carátula" in e or "Caratula" in e or "plantilla" in e.lower() for e in errors):
        if "Caratula" in sheets_missing:
            pass
        elif not errors:
            errors.append("La plantilla no contiene clientes con períodos reconocibles.")

    return EvaluacionDataset(
        source=source_label,
        sheets_read=sheets_read,
        sheets_missing=sheets_missing,
        companies=companies,
        altman_models=altman_models,
        errors=errors,
    )


def _note_optional_sheets(sheet_map: dict, sheets_read: list[str], errors: list[str]) -> None:
    """Registra presencia de Balance/Resultados sin mezclar su detalle aún."""
    for name in SHEETS_OPTIONAL:
        if name not in sheets_read or name not in sheet_map:
            continue
        try:
            ws = sheet_map[name]
            row1 = next(ws.iter_rows(min_row=1, max_row=1, max_col=120, values_only=True))
            n = sum(1 for v in row1 if isinstance(v, str) and v.strip())
            if n == 0:
                errors.append(f"{name}: hoja presente pero sin empresas en la fila de encabezado")
        except Exception:
            errors.append(f"{name}: hoja presente pero no legible (se omite)")


def _is_year(value: Any) -> bool:
    try:
        year = int(float(value))
    except (TypeError, ValueError):
        return False
    return 1990 <= year <= 2100


def _to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if not text or text.lower() in {"n/a", "na", "-", "—"}:
            return None
        try:
            return float(text)
        except ValueError:
            return None
    return None


def _norm_label(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("á", "a").replace("é", "e").replace("í", "i")
    text = text.replace("ó", "o").replace("ú", "u")
    return re.sub(r"\s+", " ", text)


def _clean_company_name(raw: str) -> tuple[str, str]:
    name = re.sub(r"\.(xlsx|xlsm|xls)$", "", raw.strip(), flags=re.I)
    name = re.sub(r"\s+v\d+\s*$", "", name, flags=re.I).strip()
    name = re.sub(r"\s+vf\s*$", "", name, flags=re.I).strip()
    code_match = re.match(r"^([A-Z]{2,4}\d*)[-_\s]+(.+)$", name)
    if code_match:
        return code_match.group(1).upper(), name
    slug = re.sub(r"[^A-Za-z0-9]+", "_", name).strip("_")[:32].upper() or "CLIENTE"
    return slug, name


def _match_metric_key(label: str) -> str | None:
    n = _norm_label(label)
    if n in CARATULA_METRIC_ALIASES:
        return CARATULA_METRIC_ALIASES[n]
    for prefix, key in CARATULA_METRIC_ALIASES.items():
        if n == prefix:
            return key
    return None


def _match_z_key(label: str) -> str | None:
    n = _norm_label(label)
    for prefix, key in Z_MODEL_ALIASES.items():
        if n.startswith(_norm_label(prefix)):
            return key
    return None


def _parse_caratula(ws, errors: list[str]) -> list[CompanyEval]:
    # Materializar filas necesarias (Carátula ~120 filas × ~80 cols).
    grid: list[list[Any]] = []
    for row in ws.iter_rows(max_row=130, max_col=120, values_only=True):
        grid.append(list(row))
    if len(grid) < 3:
        errors.append("Caratula: insuficientes filas")
        return []

    header = grid[0]
    years_row = grid[1]
    status_row = grid[2]

    # Bloques contiguos de años.
    runs: list[tuple[int, int]] = []
    i = 0
    while i < len(years_row):
        if _is_year(years_row[i]):
            j = i
            while j < len(years_row) and _is_year(years_row[j]):
                j += 1
            runs.append((i, j - 1))
            i = j
        else:
            i += 1

    companies: list[CompanyEval] = []
    for start, end in runs:
        raw_name = None
        for col in range(start, end + 1):
            cell = header[col] if col < len(header) else None
            if isinstance(cell, str) and cell.strip():
                raw_name = cell.strip()
                break
        if not raw_name:
            errors.append(f"Caratula: bloque cols {start}-{end} sin nombre")
            continue

        code, name = _clean_company_name(raw_name)
        company = CompanyEval(code=code, name=name)

        # Metadatos por bloque (columna de etiqueta en A; valor en primera col del bloque).
        company.giro = _cell_str_near(grid, "giro del negocio", start, end)
        company.currency_unit = _cell_str_near(grid, "en miles de", start, end)
        company.auditor = _cell_str_near(grid, "firma de auditoria", start, end)

        # Mapear filas de métricas / Z (solo primera sección, antes de ESTRUCTURAS %).
        metric_rows: dict[str, int] = {}
        z_rows: dict[str, int] = {}
        z_label_row: int | None = None
        for r_idx, row in enumerate(grid):
            label = row[0] if row else None
            if not label:
                continue
            n = _norm_label(label)
            if n.startswith("estructuras"):
                break
            mkey = _match_metric_key(str(label))
            if mkey and mkey not in metric_rows:
                metric_rows[mkey] = r_idx
            zkey = _match_z_key(str(label))
            if zkey and zkey not in z_rows:
                z_rows[zkey] = r_idx
            if (
                n.startswith("3. z'' para no manufactureras")
                and z_label_row is None
                and r_idx > 60
            ):
                sample = [row[c] for c in range(start, end + 1) if c < len(row)]
                if any(
                    isinstance(v, str) and v.strip().lower() in {"bien", "soso", "mal"}
                    for v in sample
                ):
                    z_label_row = r_idx

        for col in range(start, end + 1):
            year = int(float(years_row[col]))
            status = status_row[col] if col < len(status_row) else None
            if isinstance(status, str):
                status = status.strip() or None
            else:
                status = str(status) if status is not None else None

            metrics: dict[str, float | None] = {}
            for key, r_idx in metric_rows.items():
                row = grid[r_idx]
                metrics[key] = _to_float(row[col] if col < len(row) else None)

            z_scores: dict[str, float | None] = {}
            for key, r_idx in z_rows.items():
                row = grid[r_idx]
                z_scores[key] = _to_float(row[col] if col < len(row) else None)

            z_label = None
            if z_label_row is not None:
                raw = grid[z_label_row][col] if col < len(grid[z_label_row]) else None
                if isinstance(raw, str) and raw.strip():
                    z_label = raw.strip().title()
                    if z_label.lower() == "soso":
                        z_label = "Soso"

            company.periods.append(
                PeriodPoint(
                    year=year,
                    status=status,
                    metrics=metrics,
                    z_scores=z_scores,
                    z_label=z_label,
                )
            )

        if company.periods:
            companies.append(company)

    if not companies:
        errors.append("Caratula: no se detectaron empresas")
    return companies


def _cell_str_near(grid: list[list[Any]], label_prefix: str, start: int, end: int) -> str | None:
    target = _norm_label(label_prefix)
    for row in grid[:20]:
        if not row:
            continue
        if not _norm_label(row[0]).startswith(target):
            continue
        for col in range(start, end + 1):
            val = row[col] if col < len(row) else None
            if isinstance(val, str) and val.strip():
                return val.strip()
            if val is not None and not isinstance(val, str):
                return str(val)
        # A veces el valor está en una sola celda del bloque (nombre merge-like).
        for col in range(max(0, start - 1), min(len(row), end + 2)):
            val = row[col]
            if isinstance(val, str) and val.strip() and not _norm_label(val).startswith(target):
                return val.strip()
    return None


def _parse_altman(ws, errors: list[str]) -> list[AltmanModel]:
    models: list[AltmanModel] = []
    for row in ws.iter_rows(min_row=4, max_row=12, max_col=12, values_only=True):
        vals = list(row)
        if not vals or vals[0] is None:
            continue
        try:
            mid = int(float(vals[0]))
        except (TypeError, ValueError):
            continue
        name = str(vals[1]).strip() if vals[1] else f"Modelo {mid}"
        models.append(
            AltmanModel(
                id=mid,
                name=name,
                low_cut=_to_float(vals[2]),
                high_cut=_to_float(vals[3]),
                bajo_riesgo=str(vals[4]).strip() if vals[4] else None,
                riesgo_moderado=str(vals[5]).strip() if vals[5] else None,
                alto_riesgo=str(vals[6]).strip() if vals[6] else None,
            )
        )
    if not models:
        errors.append("Altman: sin modelos de rangos")
    return models
