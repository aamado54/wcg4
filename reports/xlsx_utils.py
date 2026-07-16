"""Builders Excel compactos con openpyxl."""

from __future__ import annotations

from io import BytesIO
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def build_workbook(
    sheets: Sequence[dict[str, Any]],
    *,
    report_title: str,
    stamp_label: str,
) -> bytes:
    """
    sheets: list of {name, title, headers, rows}
    """
    wb = Workbook()
    # remove default if we create sheets
    default = wb.active
    wb.remove(default)

    if not sheets:
        ws = wb.create_sheet("Sin datos")
        ws["A1"] = report_title
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = stamp_label
        ws["A2"].font = Font(bold=True)
        ws["A4"] = "No hay tablas para exportar."
    else:
        for idx, sheet in enumerate(sheets):
            name = (sheet.get("name") or f"Hoja{idx+1}")[:31]
            ws = wb.create_sheet(name)
            title = sheet.get("title") or name
            headers = list(sheet.get("headers") or [])
            rows = list(sheet.get("rows") or [])

            ws["A1"] = report_title
            ws["A1"].font = Font(bold=True, size=14)
            ws["A2"] = stamp_label
            ws["A2"].font = Font(bold=True)
            ws["A4"] = title
            ws["A4"].font = Font(bold=True, size=12)

            start_row = 6
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=start_row, column=col, value=header)
                cell.font = Font(bold=True)
            for r_i, row in enumerate(rows, start=start_row + 1):
                for c_i, value in enumerate(row, start=1):
                    ws.cell(row=r_i, column=c_i, value=value)

            for col in range(1, max(len(headers), 1) + 1):
                letter = get_column_letter(col)
                maxlen = len(str(headers[col - 1])) if headers and col <= len(headers) else 10
                for row in rows:
                    if col - 1 < len(row) and col <= 40:
                        maxlen = max(
                            maxlen,
                            min(len(str(row[col - 1] if row[col - 1] is not None else "")), 60),
                        )
                ws.column_dimensions[letter].width = min(maxlen + 2, 42)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
