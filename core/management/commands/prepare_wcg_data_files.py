"""
Genera archivos de muestra en data/wcg/ con los nombres esperados por import_wcg_data.
Útil cuando los archivos reales aún no están en el repositorio.
"""

import csv
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand
from openpyxl import Workbook


class Command(BaseCommand):
    help = "Crea archivos CSV/XLSX de muestra en data/wcg/"

    def handle(self, *args, **options):
        target = settings.BASE_DIR / "data" / "wcg"
        target.mkdir(parents=True, exist_ok=True)

        crm_path = target / "crm datos - InfoClientesWCG para CRM.csv"
        with crm_path.open("w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow([
                "Codigo", "NIT", "Nombre", "UNE", "Tipo", "Telefono", "Email",
                "Contacto", "Cargo", "Notas",
            ])
            w.writerow([
                "9852115", "9852115", "VICENTE SOLER MUNGUÍA",
                "INVESTMENT - WC LEASING", "Cliente", "5025550101",
                "contacto@soler.gt", "María Soler", "Gerente", "Import demo",
            ])
            w.writerow([
                "DEMO002", "1234567-8", "Distribuidora Me Llega, S.A.",
                "INVESTMENT - WC FACTORING", "Cliente", "5025550202",
                "jperez@melega.gt", "Juan Pérez", "CFO", "",
            ])

        risk_path = target / (
            "balon datos - Ejemplo de datos Riesgo al 31-mayo para una operacion - Base de datos Leasing.xlsx"
        )
        wb = Workbook()
        ws = wb.active
        ws.title = "Base de datos Leasing"
        ws.append([
            "Cliente", "NIT", "Operacion", "Unidad", "Saldo", "Dias_Mora",
            "Monto_Exigible", "Fecha_Snapshot", "Nivel_Riesgo", "Alerta",
        ])
        ws.append([
            "VICENTE SOLER MUNGUÍA", "9852115", "PG01260302", "LEASING",
            125000, 45, 42000, "2026-05-31", "ALTO", "SI",
        ])
        ws.append([
            "Ingenio Palo Gordo, S.A.", "8765432-1", "LG01260115", "LEASING",
            890000, 62, 120000, "2026-05-31", "CRITICO", "SI",
        ])
        wb.save(risk_path)

        pgo_cat = target / "pgo datos - Archivos para PGO.csv"
        with pgo_cat.open("w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["Archivo", "Descripcion", "Modulo"])
            w.writerow(["control tickets", "Tickets TI Q2", "PGO"])

        tickets_path = target / "pgo datos - control de tickets marzo abril y mayo 2026 para PGO.xlsx"
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Tickets"
        ws2.append([
            "Ticket", "Titulo", "Estado", "Prioridad", "Asignado", "Fecha_Apertura",
            "Fecha_Cierre", "Unidad_Negocio", "SLA_Horas",
        ])
        ws2.append([
            "TI-IMP-001", "Instalar cliente VPN", "Cerrado", "Alta", "caa",
            "2026-03-05 09:00:00", "2026-03-05 14:00:00", "TI", 48,
        ])
        ws2.append([
            "TI-IMP-002", "Error importación CRM", "Abierto", "Media", "caa",
            "2026-05-10 11:30:00", "", "TI", 48,
        ])
        wb2.save(tickets_path)

        analisis_path = target / "pgo ejemplo del analisis - PGO - TI Q22026.xlsx"
        wb3 = Workbook()
        ws3 = wb3.active
        ws3.title = "PGO TI Q2"
        ws3.append(["ID", "Asunto", "Status", "Prioridad", "Tecnico", "Creado", "Cerrado", "Area", "SLA"])
        ws3.append([
            "TI-Q2-001", "Migración servidor", "Cerrado", "Alta", "caa",
            "2026-04-01 08:00:00", "2026-04-02 10:00:00", "TI", 48,
        ])
        wb3.save(analisis_path)

        self.stdout.write(self.style.SUCCESS(f"Archivos creados en {target}"))
