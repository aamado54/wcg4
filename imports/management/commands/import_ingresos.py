# imports/management/commands/import_ingresos.py

from decimal import Decimal
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from core.models import UNE, MetricDefinition
from pgc.models import PGCPlan, MonthlyTarget, MonthlyMetricResult, MonthlyExchangeRate


class Command(BaseCommand):
    help = (
        "Importa ingresos mensuales desde archivo de estado de resultados "
        "(xlsx) y actualiza la métrica INGRESOS por UNE (no incluye INVESTMENT)."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--path",
            type=str,
            required=True,
            help="Ruta al archivo de estado de resultados (xlsx).",
        )
        parser.add_argument(
            "--year",
            type=int,
            required=True,
            help="Año del período, por ejemplo 2026.",
        )
        parser.add_argument(
            "--month",
            type=int,
            required=True,
            help="Mes del período, 1-12.",
        )

    def _infer_une_from_filename(self, path: Path) -> UNE:
        name = path.name.upper().strip()

        if name.startswith("WCI"):
            code = "INSURANCE"
        elif name.startswith("WCL"):
            code = "LEASING"
        elif name.startswith("WCF"):
            code = "FACTORING"
        elif name.startswith("WC"):
            code = "FACTORING"
        else:
            raise CommandError(
                f"No se pudo inferir UNE desde el nombre de archivo: {path.name}. "
                "Esperaba prefijos WC / WCF / WCI / WCL."
            )

        try:
            return UNE.objects.get(code=code)
        except UNE.DoesNotExist:
            raise CommandError(f"UNE con code={code} no existe en la base de datos.")

    def _to_decimal(self, raw_value, context: str) -> Decimal:
        if raw_value in (None, ""):
            return Decimal("0")

        text = str(raw_value).strip().replace(",", "")
        try:
            return Decimal(text)
        except Exception:
            raise CommandError(
                f"No se pudo convertir a decimal ({context}). Valor bruto={raw_value!r}"
            )

    def _sum_accounts_starting_with(self, ws, month: int, prefix: str) -> Decimal:
        """
        Suma cuentas contables que empiezan con `prefix` (ej: '4' o '8').
        
        Lógica:
        1) Buscar primero si existe una fila con CUENTA == prefix.
           - Si existe, usar SALDOFIN y terminar.
        2) Solo si NO existe CUENTA == prefix, usar la lógica alterna:
           - sumar filas cuyo código empiece con prefix
           - y tenga exactamente 9 dígitos
           - tomando el valor de la columna del mes en español.
        """
        cuenta_col = 2  # B

        # ---- Paso 1: detectar si existe CUENTA == prefix ----
        row_with_prefix = None
        for row in ws.iter_rows(min_row=2):
            cuenta = row[cuenta_col - 1].value
            if cuenta is None:
                continue
            if str(cuenta).strip() == prefix:
                row_with_prefix = row
                break

        # ---- Si existe CUENTA == prefix, usar SOLO esa lógica ----
        if row_with_prefix is not None:
            saldo_fin_col = None
            for col_idx, cell in enumerate(ws[1], start=1):
                if cell.value is None:
                    continue
                if str(cell.value).strip().upper() == "SALDOFIN":
                    saldo_fin_col = col_idx
                    break

            if saldo_fin_col is None:
                # No hay columna SALDOFIN, retornar cero (sin error)
                return Decimal("0")

            saldo_value = row_with_prefix[saldo_fin_col - 1].value
            return self._to_decimal(
                saldo_value,
                f"hoja Datos fila CUENTA={prefix} SALDOFIN",
            )

        # ---- Paso 2: solo si NO existe CUENTA == prefix, usar lógica mensual ----
        month_names = {
            1: "ENERO",
            2: "FEBRERO",
            3: "MARZO",
            4: "ABRIL",
            5: "MAYO",
            6: "JUNIO",
            7: "JULIO",
            8: "AGOSTO",
            9: "SEPTIEMBRE",
            10: "OCTUBRE",
            11: "NOVIEMBRE",
            12: "DICIEMBRE",
        }

        month_name = month_names.get(month)
        if not month_name:
            raise CommandError(f"Mes inválido: {month}")

        month_col = None
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value is None:
                continue
            if str(cell.value).strip().upper() == month_name:
                month_col = col_idx
                break

        if month_col is None:
            # No hay columna del mes, retornar cero (sin error)
            return Decimal("0")

        total = Decimal("0")
        for row in ws.iter_rows(min_row=2):
            cuenta = row[cuenta_col - 1].value
            if cuenta is None:
                continue

            cuenta_str = str(cuenta).strip()
            cuenta_digits = "".join(ch for ch in cuenta_str if ch.isdigit())

            if cuenta_digits.startswith(prefix) and len(cuenta_digits) == 9:
                month_value = row[month_col - 1].value
                if month_value in (None, ""):
                    continue

                total += self._to_decimal(
                    month_value,
                    f"cuenta {cuenta_str} columna {month_name}",
                )

        return total

    def _read_ingreso_from_excel(self, path: Path, month: int) -> Decimal:
        """
        Lee el ingreso bruto desde el estado de resultados.
        
        Regla fija: suma los ingresos de cuentas que empiezan con 4
        MÁS los ingresos de cuentas que empiezan con 8.
        """
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception as e:
            raise CommandError(f"No se pudo abrir el archivo Excel: {e}")

        if "Datos" not in wb.sheetnames:
            raise CommandError(
                f"El archivo {path.name} no contiene hoja 'Datos'. "
                f"Hojas disponibles: {', '.join(wb.sheetnames)}"
            )

        ws = wb["Datos"]

        # Suma cuentas que empiezan con 4
        suma_4 = self._sum_accounts_starting_with(ws, month, "4")
        
        # Suma cuentas que empiezan con 8
        suma_8 = self._sum_accounts_starting_with(ws, month, "8")

        # Total = 4 + 8
        total = suma_4 + suma_8

        self.stdout.write(
            self.style.WARNING(
                f"  Cuentas '4': {suma_4} | Cuentas '8': {suma_8} | Total: {total}"
            )
        )

        return total

    def _get_exchange_rate(self, year: int, month: int) -> Decimal:
        try:
            rate = MonthlyExchangeRate.objects.get(year=year, month=month)
        except MonthlyExchangeRate.DoesNotExist:
            raise CommandError(
                f"No existe tipo de cambio para {year}-{month:02d}. "
                "Debe registrar MonthlyExchangeRate antes de importar ingresos."
            )

        if rate.usd_to_gtq in (None, Decimal("0")):
            raise CommandError(
                f"Tipo de cambio inválido para {year}-{month:02d}: {rate.usd_to_gtq}"
            )

        return rate.usd_to_gtq

    @transaction.atomic
    def handle(self, *args, **options):
        path = Path(options["path"])
        year = options["year"]
        month = options["month"]

        if not path.exists():
            raise CommandError(f"Archivo no encontrado: {path}")

        self.stdout.write(
            self.style.WARNING(f"Leyendo estado de resultados desde {path} ...")
        )

        try:
            plan = PGCPlan.objects.get(year=year)
        except PGCPlan.DoesNotExist:
            raise CommandError(f"No existe PGCPlan para year={year}")

        try:
            metric = MetricDefinition.objects.get(
                code=MetricDefinition.CODE_INGRESOS
            )
        except MetricDefinition.DoesNotExist:
            raise CommandError("No existe MetricDefinition para CODE_INGRESOS")

        une = self._infer_une_from_filename(path)
        if une.code == "INVESTMENT":
            raise CommandError(
                "Este comando no debe usarse para INVESTMENT; "
                "para esa UNE use el recálculo desde NewClientImportRow."
            )

        self.stdout.write(self.style.WARNING(f"Archivo detectado para UNE={une.code}"))

        ingreso_bruto_gtq = self._read_ingreso_from_excel(path, month)
        
        if ingreso_bruto_gtq is None:
            raise CommandError(
                f"_read_ingreso_from_excel devolvió None para {path.name} "
                f"en {year}-{month:02d}."
            )

        # TERCERO del plan: Insurance también se divide entre mil
        if une.code in ("FACTORING", "LEASING", "INSURANCE"):
            ingreso_ajustado_gtq = ingreso_bruto_gtq / Decimal("1000")
        else:
            ingreso_ajustado_gtq = ingreso_bruto_gtq

        tipo_cambio = self._get_exchange_rate(year, month)
        ingreso_usd = ingreso_ajustado_gtq / tipo_cambio

        try:
            target = MonthlyTarget.objects.get(
                plan=plan,
                une=une,
                metric=metric,
                year=year,
                month=month,
            )
        except MonthlyTarget.DoesNotExist:
            raise CommandError(
                f"No existe MonthlyTarget para INGRESOS {une.code} "
                f"{year}-{month:02d}."
            )

        mmr, _ = MonthlyMetricResult.objects.get_or_create(
            plan=plan,
            une=une,
            metric=metric,
            year=year,
            month=month,
            defaults={"target_value": target.target_value},
        )

        mmr.target_value = target.target_value
        mmr.measured_value = ingreso_usd

        measured = mmr.measured_value or Decimal("0")
        target_val = mmr.target_value or Decimal("0")
        achieved = measured >= target_val

        mmr.is_achieved = achieved
        mmr.points_awarded = target.points_if_achieved if achieved else 0
        mmr.calculation_note = (
            f"Ingreso importado desde {path.name}, hoja Datos. "
            f"BrutoGTQ(4+8)={ingreso_bruto_gtq}, "
            f"AjustadoGTQ={ingreso_ajustado_gtq}, "
            f"TipoCambioGTQxUSD={tipo_cambio}, "
            f"USD={ingreso_usd}, UNE={une.code}."
        )
        mmr.save()

        self.stdout.write(
            self.style.SUCCESS(
                f"Actualizado INGRESOS {une.code} {year}-{month:02d}: "
                f"real={measured} meta={target_val} "
                f"logrado={achieved} puntos={mmr.points_awarded}"
            )
        )
        self.stdout.write(self.style.SUCCESS("Import INGRESOS completado."))